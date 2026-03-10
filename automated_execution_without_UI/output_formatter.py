import json
import os
from dataclasses import asdict
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import (
    TextModerationResult,
    ImageModerationResult,
    VideoModerationResult,
)


# ---------------------------------------------------------------------------
# JSON output
# ---------------------------------------------------------------------------

def save_results_json(text_results, image_results, video_results, model_id, output_path):
    """Write results.json containing all moderation results."""
    data = {
        "run_timestamp": datetime.now(timezone.utc).isoformat(),
        "model_id": model_id,
        "text_moderation": [_serialize_result(r) for r in text_results],
        "image_moderation": [_serialize_result(r) for r in image_results],
        "video_moderation": [_serialize_result(r) for r in video_results],
        "timing_summary": compute_timing_summary(text_results, image_results, video_results),
    }
    filepath = os.path.join(output_path, "results.json")
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    return filepath


def _serialize_result(result):
    """Convert a frozen dataclass to a JSON-safe dict."""
    d = asdict(result)
    # Remove raw bytes that can't be serialized
    return d


# ---------------------------------------------------------------------------
# Summary text
# ---------------------------------------------------------------------------

def save_summary_txt(text_results, image_results, video_results, model_id, output_path):
    """Write summary.txt with human-readable tables."""
    timestamp = datetime.now(timezone.utc).isoformat()
    lines = [
        f"Run: {timestamp}",
        f"Model: {model_id}",
        "",
    ]

    # --- TEXT ---
    lines.append("=" * 60)
    lines.append(f"  TEXT MODERATION ({len(text_results)} rows)")
    lines.append("=" * 60)
    lines.append(
        f" {'Row':>3} | {'Time(s)':>8} | {'Risk':>8} | {'Porn':>6} | {'Violence':>8} | {'Tobacco':>7} | {'Political':>9} | {'Profanity':>9} | Summary"
    )
    lines.append("-" * 125)
    for r in text_results:
        if r.error:
            lines.append(f" {r.row_index:>3} | {'ERROR':>8} | {'-':>8} | {'-':>6} | {'-':>8} | {'-':>7} | {'-':>9} | {'-':>9} | {r.error[:50]}")
        else:
            m = r.moderation
            risk = m.overall_risk if m else "-"
            porn = m.pornography.severity if m else "-"
            viol = m.violence.severity if m else "-"
            toba = m.tobacco_alcohol.severity if m else "-"
            poli = m.political_sensitivity.severity if m else "-"
            prof = m.profanity.severity if m else "-"
            summ = (m.summary if m else "")[:50]
            lines.append(
                f" {r.row_index:>3} | {r.moderation_time_sec:>8.3f} | {risk:>8} | {porn:>6} | {viol:>8} | {toba:>7} | {poli:>9} | {prof:>9} | {summ}"
            )
    lines.append("")

    # --- IMAGE ---
    lines.append("=" * 60)
    lines.append(f"  IMAGE MODERATION ({len(image_results)} rows)")
    lines.append("=" * 60)
    lines.append(
        f" {'Row':>3} | {'DL(s)':>8} | {'Mod(s)':>8} | {'Size(KB)':>8} | {'Risk':>8} | {'Detected Categories':30} | Summary"
    )
    lines.append("-" * 120)
    for r in image_results:
        if r.error:
            lines.append(f" {r.row_index:>3} | {r.download_time_sec:>8.3f} | {'ERROR':>8} | {'-':>8} | {'-':>8} | {'-':30} | {r.error[:50]}")
        else:
            m = r.moderation
            size_kb = r.image_size_bytes // 1024
            risk = m.overall_risk if m else "-"
            detected = _detected_categories(m) if m else "-"
            summ = (m.summary if m else "")[:50]
            lines.append(
                f" {r.row_index:>3} | {r.download_time_sec:>8.3f} | {r.moderation_time_sec:>8.3f} | {size_kb:>8} | {risk:>8} | {detected:30} | {summ}"
            )
    lines.append("")

    # --- VIDEO ---
    lines.append("=" * 60)
    lines.append(f"  VIDEO MODERATION ({len(video_results)} rows)")
    lines.append("=" * 60)
    lines.append(
        f" {'Row':>3} | {'DL(s)':>8} | {'Mod(s)':>8} | {'Method':>10} | {'Size(MB)':>8} | {'Risk':>8} | {'Detected Categories':30} | Summary"
    )
    lines.append("-" * 130)
    for r in video_results:
        if r.error:
            lines.append(
                f" {r.row_index:>3} | {r.download_time_sec:>8.3f} | {'ERROR':>8} | {r.analysis_method:>10} | {'-':>8} | {'-':>8} | {'-':30} | {r.error[:40]}"
            )
        else:
            m = r.moderation
            size_mb = r.video_size_bytes / (1024 * 1024)
            risk = m.overall_risk if m else "-"
            detected = _detected_categories(m) if m else "-"
            summ = (m.summary if m else "")[:40]
            lines.append(
                f" {r.row_index:>3} | {r.download_time_sec:>8.3f} | {r.moderation_time_sec:>8.3f} | {r.analysis_method:>10} | {size_mb:>8.1f} | {risk:>8} | {detected:30} | {summ}"
            )
    lines.append("")

    # --- TIMING ---
    timing = compute_timing_summary(text_results, image_results, video_results)
    lines.append("=" * 60)
    lines.append("  TIMING SUMMARY")
    lines.append("=" * 60)
    lines.append(f" {'Category':>9} | {'Count':>5} | {'Success':>7} | {'Avg API(s)':>10} | {'Avg Total(s)':>12}")
    lines.append("-" * 55)
    for cat in ("text", "image", "video"):
        t = timing.get(cat, {})
        lines.append(
            f" {cat.capitalize():>9} | {t.get('count', 0):>5} | {t.get('success', 0):>7} | "
            f"{t.get('avg_api_sec', 0):>10.3f} | {t.get('avg_total_sec', 0):>12.3f}"
        )
    lines.append("")

    filepath = os.path.join(output_path, "summary.txt")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return filepath


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

_CATEGORY_NAMES = ("pornography", "violence", "tobacco_alcohol", "political_sensitivity", "profanity")

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_SEVERITY_FILLS = {
    "high": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "medium": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "low": PatternFill(start_color="B5E48C", end_color="B5E48C", fill_type="solid"),
    "critical": PatternFill(start_color="D90429", end_color="D90429", fill_type="solid"),
}
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def save_results_xlsx(text_results, image_results, video_results, model_id, output_path):
    """Write results.xlsx with 4 sheets: Text, Image, Video, Summary."""
    wb = openpyxl.Workbook()

    _build_text_sheet(wb, text_results)
    _build_image_sheet(wb, image_results)
    _build_video_sheet(wb, video_results)
    _build_summary_sheet(wb, text_results, image_results, video_results, model_id)

    # Remove the default empty sheet created by Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    filepath = os.path.join(output_path, "results.xlsx")
    wb.save(filepath)
    return filepath


def _style_header_row(ws, num_cols):
    """Apply header styling to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_severity_fill(cell, severity):
    """Color-code a cell by severity level."""
    fill = _SEVERITY_FILLS.get(severity)
    if fill:
        if severity == "critical":
            cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill


def _auto_column_widths(ws, min_width=10, max_width=60):
    """Set column widths based on content, capped at max_width."""
    for col_idx in range(1, ws.max_column + 1):
        widest = min_width
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                widest = max(widest, min(len(str(val)), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = widest + 2


def _cat_severity(moderation, name):
    if moderation is None:
        return "-"
    return getattr(moderation, name).severity


def _cat_detected(moderation, name):
    if moderation is None:
        return False
    return getattr(moderation, name).detected


def _cat_details(moderation, name):
    if moderation is None:
        return ""
    return getattr(moderation, name).details


# --- Text sheet ---

def _build_text_sheet(wb, results):
    ws = wb.create_sheet("Text Moderation")
    headers = [
        "Row", "Moderation Time(s)", "Overall Risk",
        "Porn Severity", "Porn Details",
        "Violence Severity", "Violence Details",
        "Tobacco Severity", "Tobacco Details",
        "Political Severity", "Political Details",
        "Profanity Severity", "Profanity Details",
        "Summary", "Original Text", "Error",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in results:
        m = r.moderation
        row = [
            r.row_index,
            r.moderation_time_sec,
            m.overall_risk if m else ("ERROR" if r.error else "-"),
            _cat_severity(m, "pornography"), _cat_details(m, "pornography"),
            _cat_severity(m, "violence"), _cat_details(m, "violence"),
            _cat_severity(m, "tobacco_alcohol"), _cat_details(m, "tobacco_alcohol"),
            _cat_severity(m, "political_sensitivity"), _cat_details(m, "political_sensitivity"),
            _cat_severity(m, "profanity"), _cat_details(m, "profanity"),
            m.summary if m else "",
            r.original_text[:500],
            r.error or "",
        ]
        ws.append(row)
        row_idx = ws.max_row
        # Color severity cells
        for col, cat_name in [(4, "pornography"), (6, "violence"), (8, "tobacco_alcohol"),
                              (10, "political_sensitivity"), (12, "profanity")]:
            _apply_severity_fill(ws.cell(row=row_idx, column=col), _cat_severity(m, cat_name))
        # Color overall risk
        _apply_severity_fill(ws.cell(row=row_idx, column=3), m.overall_risk if m else "")
        # Error row highlight
        if r.error:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _ERROR_FILL
        # Wrap text for details and summary columns
        for col in (5, 7, 9, 11, 13, 14, 15):
            ws.cell(row=row_idx, column=col).alignment = _WRAP_ALIGNMENT

    _auto_column_widths(ws)


# --- Image sheet ---

def _build_image_sheet(wb, results):
    ws = wb.create_sheet("Image Moderation")
    headers = [
        "Row", "Download Time(s)", "Moderation Time(s)", "Size(KB)", "Overall Risk",
        "Porn Severity", "Porn Details",
        "Violence Severity", "Violence Details",
        "Tobacco Severity", "Tobacco Details",
        "Political Severity", "Political Details",
        "Profanity Severity", "Profanity Details",
        "Summary", "Image URL", "Error",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in results:
        m = r.moderation
        row = [
            r.row_index,
            r.download_time_sec,
            r.moderation_time_sec,
            r.image_size_bytes // 1024,
            m.overall_risk if m else ("ERROR" if r.error else "-"),
            _cat_severity(m, "pornography"), _cat_details(m, "pornography"),
            _cat_severity(m, "violence"), _cat_details(m, "violence"),
            _cat_severity(m, "tobacco_alcohol"), _cat_details(m, "tobacco_alcohol"),
            _cat_severity(m, "political_sensitivity"), _cat_details(m, "political_sensitivity"),
            _cat_severity(m, "profanity"), _cat_details(m, "profanity"),
            m.summary if m else "",
            r.image_url,
            r.error or "",
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col, cat_name in [(6, "pornography"), (8, "violence"), (10, "tobacco_alcohol"),
                              (12, "political_sensitivity"), (14, "profanity")]:
            _apply_severity_fill(ws.cell(row=row_idx, column=col), _cat_severity(m, cat_name))
        _apply_severity_fill(ws.cell(row=row_idx, column=5), m.overall_risk if m else "")
        if r.error:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _ERROR_FILL
        for col in (7, 9, 11, 13, 15, 16):
            ws.cell(row=row_idx, column=col).alignment = _WRAP_ALIGNMENT

    _auto_column_widths(ws)


# --- Video sheet ---

def _build_video_sheet(wb, results):
    ws = wb.create_sheet("Video Moderation")
    headers = [
        "Row", "Download Time(s)", "Moderation Time(s)", "Method", "Size(MB)", "Overall Risk",
        "Porn Severity", "Porn Details",
        "Violence Severity", "Violence Details",
        "Tobacco Severity", "Tobacco Details",
        "Political Severity", "Political Details",
        "Profanity Severity", "Profanity Details",
        "Summary", "Video URL", "Error",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in results:
        m = r.moderation
        row = [
            r.row_index,
            r.download_time_sec,
            r.moderation_time_sec,
            r.analysis_method,
            round(r.video_size_bytes / (1024 * 1024), 1),
            m.overall_risk if m else ("ERROR" if r.error else "-"),
            _cat_severity(m, "pornography"), _cat_details(m, "pornography"),
            _cat_severity(m, "violence"), _cat_details(m, "violence"),
            _cat_severity(m, "tobacco_alcohol"), _cat_details(m, "tobacco_alcohol"),
            _cat_severity(m, "political_sensitivity"), _cat_details(m, "political_sensitivity"),
            _cat_severity(m, "profanity"), _cat_details(m, "profanity"),
            m.summary if m else "",
            r.video_url,
            r.error or "",
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col, cat_name in [(7, "pornography"), (9, "violence"), (11, "tobacco_alcohol"),
                              (13, "political_sensitivity"), (15, "profanity")]:
            _apply_severity_fill(ws.cell(row=row_idx, column=col), _cat_severity(m, cat_name))
        _apply_severity_fill(ws.cell(row=row_idx, column=6), m.overall_risk if m else "")
        if r.error:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = _ERROR_FILL
        for col in (8, 10, 12, 14, 16, 17):
            ws.cell(row=row_idx, column=col).alignment = _WRAP_ALIGNMENT

    _auto_column_widths(ws)


# --- Summary sheet ---

def _build_summary_sheet(wb, text_results, image_results, video_results, model_id):
    ws = wb.create_sheet("Summary")
    timestamp = datetime.now(timezone.utc).isoformat()

    # Run info
    ws.append(["Run Timestamp", timestamp])
    ws.append(["Model", model_id])
    ws.append([])

    # Timing summary
    timing = compute_timing_summary(text_results, image_results, video_results)
    timing_headers = ["Category", "Count", "Success", "Errors", "Avg API(s)", "Avg Total(s)"]
    ws.append(timing_headers)
    _style_header_row_at(ws, ws.max_row, len(timing_headers))
    for cat in ("text", "image", "video"):
        t = timing.get(cat, {})
        ws.append([
            cat.capitalize(),
            t.get("count", 0),
            t.get("success", 0),
            t.get("count", 0) - t.get("success", 0),
            t.get("avg_api_sec", 0),
            t.get("avg_total_sec", 0),
        ])
    ws.append([])

    # Detection summary per category
    ws.append(["Detection Summary by Category"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append([])

    all_results = (
        [("Text", r) for r in text_results]
        + [("Image", r) for r in image_results]
        + [("Video", r) for r in video_results]
    )

    det_headers = ["Category", "Total Checked", "Detected Count", "Detection Rate",
                    "High/Critical Count", "Severity Distribution"]
    ws.append(det_headers)
    _style_header_row_at(ws, ws.max_row, len(det_headers))

    for cat_name in _CATEGORY_NAMES:
        total = 0
        detected = 0
        severity_counts = {}
        for _, r in all_results:
            if r.error or r.moderation is None:
                continue
            total += 1
            cat_obj = getattr(r.moderation, cat_name)
            if cat_obj.detected:
                detected += 1
            sev = cat_obj.severity
            severity_counts[sev] = severity_counts.get(sev, 0) + 1

        high_critical = severity_counts.get("high", 0) + severity_counts.get("critical", 0)
        dist = ", ".join(f"{k}={v}" for k, v in sorted(severity_counts.items()) if k != "none")
        rate = f"{detected / total * 100:.1f}%" if total else "N/A"
        ws.append([cat_name, total, detected, rate, high_critical, dist or "all none"])

    ws.append([])

    # Overall risk distribution
    ws.append(["Overall Risk Distribution"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append([])
    risk_headers = ["Content Type", "safe", "low", "medium", "high", "critical", "error/parse_fail"]
    ws.append(risk_headers)
    _style_header_row_at(ws, ws.max_row, len(risk_headers))

    for label, results_list in [("Text", text_results), ("Image", image_results), ("Video", video_results)]:
        counts = {"safe": 0, "low": 0, "medium": 0, "high": 0, "critical": 0, "error": 0}
        for r in results_list:
            if r.error or r.moderation is None:
                counts["error"] += 1
            else:
                risk = r.moderation.overall_risk
                counts[risk] = counts.get(risk, 0) + 1
        ws.append([label, counts["safe"], counts["low"], counts["medium"],
                    counts["high"], counts["critical"], counts["error"]])

    _auto_column_widths(ws, min_width=12)


def _style_header_row_at(ws, row_num, num_cols):
    """Apply header styling to a specific row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _detected_categories(moderation):
    """Return comma-separated list of detected category names."""
    detected = []
    for name in ("pornography", "violence", "tobacco_alcohol", "political_sensitivity", "profanity"):
        cat = getattr(moderation, name)
        if cat.detected:
            detected.append(name)
    return ", ".join(detected) if detected else "none"


def compute_timing_summary(text_results, image_results, video_results):
    """Compute per-category timing statistics."""
    summary = {}

    # Text
    success_text = [r for r in text_results if r.error is None]
    summary["text"] = {
        "count": len(text_results),
        "success": len(success_text),
        "avg_api_sec": _avg([r.moderation_time_sec for r in success_text]),
        "avg_total_sec": _avg([r.moderation_time_sec for r in success_text]),
    }

    # Image
    success_img = [r for r in image_results if r.error is None]
    summary["image"] = {
        "count": len(image_results),
        "success": len(success_img),
        "avg_api_sec": _avg([r.moderation_time_sec for r in success_img]),
        "avg_total_sec": _avg([r.download_time_sec + r.moderation_time_sec for r in success_img]),
    }

    # Video
    success_vid = [r for r in video_results if r.error is None]
    summary["video"] = {
        "count": len(video_results),
        "success": len(success_vid),
        "avg_api_sec": _avg([r.moderation_time_sec for r in success_vid]),
        "avg_total_sec": _avg([r.download_time_sec + r.moderation_time_sec for r in success_vid]),
    }

    return summary


def _avg(values):
    return round(sum(values) / len(values), 3) if values else 0.0
